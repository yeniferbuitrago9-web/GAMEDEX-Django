# Módulo principal de vistas para GAMEDEX (usuarios, carrito, productos, facturación)
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth import logout
from django.views.decorators.cache import never_cache
from django.http import HttpResponse, HttpResponseForbidden
from reportlab.pdfgen import canvas
from django.http import JsonResponse
from requests import post
from GAMEDEX.forms import EditarPerfilForm
from .models import Comentario, Comunidad, Perfil, Producto, Publicacion
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from .forms import ComunidadForm
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import get_template
from django.http import HttpResponse
from django.contrib.auth import update_session_auth_hash
from datetime import timedelta
from django.utils import timezone



# =====================================
# INICIO
# =====================================


def usuario_online(perfil):
    if not perfil.last_seen:
        return False

    return timezone.now() - perfil.last_seen < timedelta(minutes=2)

def api_productos(request):
    productos = Producto.objects.all()[:20]

    data = [
        {
            "id": p.id,
            "nombre": p.nombre,
            "precio": str(p.precio)
        }
        for p in productos
    ]

    return JsonResponse(data, safe=False)

def inicio(request):
    print("ENTRANDO A INICIO 🔥")

    productos = Producto.objects.filter(publicado=True, destacado=True)
    print("PRODUCTOS:", productos)

    return render(request, 'inicio.html', {'productos': productos})

# =====================================
# REGISTRO (MÉTODO SEGURO TRADICIONAL)
# =====================================
def registro(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)   
        if form.is_valid():
            user = form.save()

            grupo, _ = Group.objects.get_or_create(name='Usuario')
            user.groups.add(grupo)

            messages.success(request, "Usuario registrado correctamente.")
            return redirect('login')
        else:
            # Si el formulario tiene errores, se los mandamos a los mensajes
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
    else:
        form = UserCreationForm()

    return render(request, 'registro.html', {'form': form})


# =====================================
# REGISTRO PUBLICO (MÉTODO PARA EL HTML CIBER)
# =====================================
def registro_publico(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        rol = request.POST.get("rol")

        # 1. Validaciones previas obligatorias
        if not username or not password:
            messages.error(request, "El nombre de usuario y la contraseña son obligatorios.")
            return render(request, "registro_publico.html")

        if User.objects.filter(username=username).exists():
            messages.error(request, "El nombre de usuario ya se encuentra registrado.")
            return render(request, "registro_publico.html")
            
        if email and User.objects.filter(email=email).exists():
            messages.error(request, "Este correo electrónico ya está en uso.")
            return render(request, "registro_publico.html")

        # 2. Intentar la creación segura del usuario
        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password
            )

            # Asignar grupo según elección de tu select
            if rol == "Vendedor":
                grupo, _ = Group.objects.get_or_create(name="Vendedor")
                rol_perfil = "Vendedor"
            else:
                grupo, _ = Group.objects.get_or_create(name="Cliente")
                rol_perfil = "Cliente"

            user.groups.add(grupo)

            # Sincronizar el rol en el perfil (el signal lo crea con 'Usuario' por defecto)
            try:
                user.perfil.rol = rol_perfil
                user.perfil.save()
            except Exception:
                pass

            messages.success(request, "¡Cuenta creada correctamente! Ya puedes iniciar sesión.")
            return redirect("login")

        except Exception as e:
            messages.error(request, f"Error inesperado al crear la cuenta: {str(e)}")
            return render(request, "registro_publico.html")

    return render(request, "registro_publico.html")
# =====================================
# DASHBOARD USUARIO
# =====================================
@login_required
@never_cache
def dashboard_usuario(request):

    productos = Producto.objects.filter(publicado=True)

    clave = get_carrito_key(request)
    carrito = request.session.get(clave, {})

    # Si la sesión está vacía pero hay carrito guardado en BD, restaurarlo
    if not carrito:
        try:
            guardado = request.user.perfil.carrito_guardado
            if guardado and guardado != "{}":
                carrito = json.loads(guardado)
                request.session[clave] = carrito
                request.session.modified = True
        except Exception:
            pass

    total_carrito = sum(item["cantidad"] for item in carrito.values())


    return render(request, "dashboard_usuario.html", {
        "productos": productos,
        "total_carrito": total_carrito
    })


# =====================================
# CARRITO
# =====================================
def get_carrito_key(request):
    """Clave de sesión única por usuario para que el carrito persista entre sesiones."""
    if request.user.is_authenticated:
        return f"carrito_{request.user.id}"
    return "carrito"

@login_required
def agregar_carrito(request, producto_id):

    producto = get_object_or_404(Producto, id=producto_id)

    try:
        cantidad = int(request.POST.get("cantidad", 1))
        if cantidad < 1:
            cantidad = 1
    except (ValueError, TypeError):
        cantidad = 1

    clave = get_carrito_key(request)
    carrito = request.session.get(clave, {})

    if str(producto_id) in carrito:
        carrito[str(producto_id)]["cantidad"] += cantidad
    else:
        carrito[str(producto_id)] = {
            "nombre": producto.nombre,
            "precio": float(producto.precio),
            "cantidad": cantidad
        }

    # No superar el stock disponible
    if carrito[str(producto_id)]["cantidad"] > producto.cantidad:
        carrito[str(producto_id)]["cantidad"] = producto.cantidad

    request.session[clave] = carrito
    request.session.modified = True

    return redirect("dashboard_usuario")

def ver_carrito(request):

    clave = get_carrito_key(request)
    carrito = request.session.get(clave, {})
    productos_carrito = []
    total = 0
    hubo_cambios = False

    for producto_id in list(carrito.keys()):
        item = carrito[producto_id]
        producto = Producto.objects.filter(id=producto_id).first()

        if not producto:
            # El producto ya no existe (fue eliminado del catálogo), lo quitamos del carrito
            del carrito[producto_id]
            hubo_cambios = True
            continue

        cantidad = item["cantidad"]
        precio = float(item["precio"])

        subtotal = precio * cantidad
        total += subtotal

        productos_carrito.append({
            "producto": producto,
            "cantidad": cantidad,
            "subtotal": subtotal
        })

    if hubo_cambios:
        request.session[clave] = carrito
        request.session.modified = True

    return render(request, "carrito.html", {
        "productos_carrito": productos_carrito,
        "total": total
    })


def quitar_unidad(request, producto_id):

    clave = get_carrito_key(request)
    carrito = request.session.get(clave, {})

    if str(producto_id) in carrito:
        carrito[str(producto_id)]["cantidad"] -= 1

        if carrito[str(producto_id)]["cantidad"] <= 0:
            del carrito[str(producto_id)]

    request.session[clave] = carrito
    request.session.modified = True

    return redirect("ver_carrito")

def eliminar_producto(request, producto_id):

    clave = get_carrito_key(request)
    carrito = request.session.get(clave, {})

    if str(producto_id) in carrito:
        del carrito[str(producto_id)]

    request.session[clave] = carrito
    request.session.modified = True

    return redirect("ver_carrito")


# =====================================
# COMPRA
# =====================================

def comprar_carrito(request):

    clave = get_carrito_key(request)
    carrito = request.session.get(clave, {})

    if not carrito:
        messages.error(request, "El carrito está vacío.")
        return redirect("dashboard_usuario")

    productos_factura = []
    total = 0

    for producto_id, item in carrito.items():

        producto = get_object_or_404(Producto, id=producto_id)

        cantidad = item.get("cantidad", 0)
        precio = float(item.get("precio", 0))

        if cantidad > producto.cantidad:
            messages.error(request, f"No hay suficiente stock de {producto.nombre}")
            return redirect("ver_carrito")

        subtotal = precio * cantidad

        producto.cantidad -= cantidad
        producto.save()

        productos_factura.append({
            "nombre": producto.nombre,
            "descripcion": producto.descripcion,
            "precio": precio,
            "cantidad": cantidad,
            "subtotal": subtotal,

        })

        total += subtotal

    request.session["factura"] = {
        "productos": productos_factura,
        "total": total
    }

    request.session[get_carrito_key(request)] = {}

    request.session.modified = True
    request.session.save()

    return redirect("/factura/")

    # =====================================
    # 🔥 GUARDAR FACTURA (FORMA SEGURA)
    # =====================================
    request.session["factura"] = {
        "productos": productos_factura,
        "total": total
    }

    # limpiar carrito
    request.session[get_carrito_key(request)] = {}

    # 🔥 FORZAR GUARDADO REAL EN SESIÓN
    request.session.modified = True
    request.session.save()

    return redirect("factura")  

# =====================================
# FACTURA
# =====================================
def factura(request):

    factura = request.session.get("factura")

    if not factura:
        messages.error(request, "No hay factura disponible.")
        return redirect("dashboard_usuario")

    return render(request, "factura.html", {
        "productos": factura["productos"],
        "total": factura["total"],
        "usuario": request.user
    })


def descargar_factura_pdf(request):

    factura = request.session.get("factura")

    if not factura:
        return HttpResponse("No hay factura disponible")

    template = get_template("descargar_factura_pdf.html")

    html = template.render({
        "productos": factura["productos"],
        "total": factura["total"],
        "usuario": request.user
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="factura_gamedex.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Error al generar PDF")

    return response

# =====================================
# PERFIL
# =====================================

@login_required
def editar_perfil_usuario(request):
    user = request.user
    perfil = user.perfil

    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')
        perfil.telefono = request.POST.get('telefono')
        perfil.direccion = request.POST.get('direccion')

        nueva_password = request.POST.get('password', '').strip()
        if nueva_password:
            user.set_password(nueva_password)

        user.save()
        perfil.save()

        update_session_auth_hash(request, user)

        messages.success(request, "Perfil actualizado correctamente")
        return redirect('dashboard_usuario')

    return render(request, 'editar_perfil.html', {
        'user': user,
        'perfil': perfil,
        'url_cancelar': 'dashboard_usuario' # 👈 Destino del botón cancelar
    })


@login_required
def editar_perfil_vendedor(request):
    user = request.user 
    perfil = user.perfil 

    if request.method == 'POST':
        form = EditarPerfilForm(request.POST, instance=user)

        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data.get("password")

            if password:
                user.set_password(password)  
            
            # Guardamos los datos del usuario (encriptados o comunes)
            user.save()

            # Guardamos los datos de los inputs manuales de tu perfil
            perfil.telefono = request.POST.get('telefono')
            perfil.direccion = request.POST.get('direccion')
            perfil.save()

            # 🔥 REFRESCAR LA SESIÓN (DEBE IR AQUÍ FUERA)
            # No importa si cambió la contraseña o solo el username/email,
            # esto le dice a Django que mantenga al vendedor logueado.
            update_session_auth_hash(request, user)

            messages.success(request, "Perfil de vendedor actualizado correctamente")
            return redirect("dashboard_vendedor")

    else:
        form = EditarPerfilForm(instance=user)

    return render(request, "editar_perfil.html", {
        "form": form,
        "user": user,
        "perfil": perfil,
        "url_cancelar": "dashboard_vendedor"
    })

# =====================================
# DASHBOARD ADMIN
# =====================================
@login_required
@never_cache
def dashboard_admin(request):

    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("redireccion_dashboard")

    # Crear perfiles faltantes para usuarios que no tienen uno
    for user in User.objects.all():
        perfil, created = Perfil.objects.get_or_create(user=user)
        if created:
            # Asignar rol según su grupo
            if user.groups.filter(name="Administrador").exists():
                perfil.rol = "Administrador"
            elif user.groups.filter(name="Vendedor").exists():
                perfil.rol = "Vendedor"
            else:
                perfil.rol = "Usuario"
            perfil.save()

    perfiles = Perfil.objects.select_related("user").all().order_by("user__username")

    # CONTADORES
    total_usuarios = Perfil.objects.count()
    total_vendedores = Perfil.objects.filter(rol="Vendedor").count()
    total_admins = Perfil.objects.filter(rol="Administrador").count()

    return render(request, "dashboard_admin.html", {
        "perfiles": perfiles,
        "total_usuarios": total_usuarios,
        "total_vendedores": total_vendedores,
        "total_admins": total_admins
    })

def lista_usuarios(request):
    query = request.GET.get("q", "")
    rol = request.GET.get("rol", "")

    perfiles = Perfil.objects.select_related("user").all()

    if query:
        perfiles = perfiles.filter(
            Q(user__username__icontains=query) |
            Q(user__email__icontains=query)
        )

    if rol:
        perfiles = perfiles.filter(rol=rol)

    # 🔥 AGREGAMOS ESTADO ONLINE
    for p in perfiles:
        p.online = usuario_online(p)

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "partials/tabla_usuarios.html", {
            "perfiles": perfiles
        })

    return render(request, "admin/usuarios.html", {
        "perfiles": perfiles,
        "total_usuarios": Perfil.objects.count(),
        "total_vendedores": Perfil.objects.filter(rol="Vendedor").count(),
        "total_admins": Perfil.objects.filter(rol="Administrador").count(),
    })

# =====================================
# COMUNIDAD Y PUBLICACIONES
# =====================================

def admin_comunidades(request):
    if not request.user.groups.filter(name="Administrador").exists():
        return redirect('inicio')

    comunidades = Comunidad.objects.all()

    return render(request, 'admin_comunidades.html', {
        'comunidades': comunidades
    })

def ver_comunidad(request, id):
    comunidad = get_object_or_404(Comunidad, id=id)
    publicaciones = Publicacion.objects.filter(comunidad=comunidad).order_by('-id')

    if request.method == "POST":
        contenido = request.POST.get("contenido")
        imagen = request.FILES.get("imagen")
        if contenido or imagen:
            Publicacion.objects.create(
                contenido=contenido,
                imagen=imagen,
                autor=request.user,
                comunidad=comunidad
            )
        return redirect('ver_comunidad', id=comunidad.id)

    return render(request, "ver_comunidad.html", {
        "comunidad": comunidad,
        "publicaciones": publicaciones
    })

def crear_publicacion(request, id) -> JsonResponse:
    if request.method == "POST" and request.user.is_authenticated:
        comunidad = get_object_or_404(Comunidad, id=id)
        contenido = request.POST.get("contenido", "").strip()
        imagen = request.FILES.get("imagen", None)

        if contenido or imagen:
            publicacion = Publicacion.objects.create(
                comunidad=comunidad,
                autor=request.user,
                contenido=contenido,
                imagen=imagen
            )

            # Retornamos datos de la publicación como JSON
            data = {
                "id": publicacion.id,
                "autor": publicacion.autor.username,
                "contenido": publicacion.contenido,
                "imagen_url": publicacion.imagen.url if publicacion.imagen else "",
                "fecha": publicacion.created_at.strftime("%Y-%m-%d %H:%M"),
            }
            return JsonResponse(data)
        
    return JsonResponse({"error": "No se pudo crear la publicación"}, status=400)

def eliminar_publicacion(request, post_id):
    if request.method == "POST":
        post = get_object_or_404(Publicacion, id=post_id)
        user = request.user

        # Verificar permisos
        if user == post.autor or user == post.comunidad.creador or getattr(user.perfil, 'rol', '') == "Administrador":
            post.delete()
            return JsonResponse({"success": True})
        else:
            return JsonResponse({"success": False, "error": "No tienes permisos para eliminar esta publicación"})

    return JsonResponse({"success": False, "error": "Método no permitido"})


@login_required
def crear_publicacion_ajax(request, id):
    if request.method == "POST":
        comunidad = get_object_or_404(Comunidad, id=id)
        contenido = request.POST.get("contenido")
        imagen = request.FILES.get("imagen")
        if not contenido and not imagen:
            return JsonResponse({"error": "Debes escribir algo o subir una imagen."})
        
        publicacion = Publicacion.objects.create(
            contenido=contenido,
            imagen=imagen,
            autor=request.user,
            comunidad=comunidad
        )

        return JsonResponse({
            "id": publicacion.id,
            "contenido": publicacion.contenido,
            "autor": publicacion.autor.username,
            "imagen_url": publicacion.imagen.url if publicacion.imagen else ""
        })
    return JsonResponse({"error": "Método no permitido."})

def dar_like(request, id):
    post = get_object_or_404(Publicacion, id=id)

    if request.user in post.likes.all():
        post.likes.remove(request.user)
        liked = False
    else:
        post.likes.add(request.user)
        liked = True

    return JsonResponse({
        "liked": liked,
        "total_likes": post.total_likes()
    })

def comentar(request, id):
    post = get_object_or_404(Publicacion, id=id)

    if request.method == "POST":
        contenido = request.POST.get("contenido")

        Comentario.objects.create(
            publicacion=post,
            autor=request.user,
            contenido=contenido
        )

    return redirect('ver_comunidad', id=post.comunidad.id)

def lista_comunidades(request):
    comunidades = Comunidad.objects.all()

    es_admin = request.user.is_authenticated and request.user.groups.filter(name="Administrador").exists()

    return render(request, "comunidades.html", {
        "comunidades": comunidades,
        "es_admin": es_admin
    })

def crear_comunidad(request):
    if request.method == "POST":

        print("FILES:", request.FILES)  # 👈 AQUÍ
        print("POST:", request.POST)    # (opcional, para ver todo)

        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")
        imagen = request.FILES.get("imagen")

        print("Imagen recibida:", imagen)  # 👈 EXTRA DEBUG

        Comunidad.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            imagen=imagen,
            creador=request.user
        )

        return redirect('comunidades')

    return render(request, 'crear_comunidad.html')

def editar_comunidad(request, id):
    comunidad = get_object_or_404(Comunidad, id=id)

    # MODIFICACIÓN: Usamos el sistema de grupos de Django para mayor estabilidad
    es_admin = request.user.groups.filter(name="Administrador").exists()
    
    # Verifica permisos: creador o admin
    if request.user != comunidad.creador and not es_admin:
        return HttpResponseForbidden("No tienes permiso para editar esta comunidad.")

    if request.method == "POST":
        form = ComunidadForm(request.POST, request.FILES, instance=comunidad)
        if form.is_valid():
            form.save()
            return redirect('ver_comunidad', id=comunidad.id)
    else:
        form = ComunidadForm(instance=comunidad)

    return render(request, "editar_comunidad.html", {"form": form, "comunidad": comunidad})

def eliminar_comunidad(request, id):
    comunidad = get_object_or_404(Comunidad, id=id)

    # MODIFICACIÓN: Validamos igual que en editar_comunidad
    es_admin = request.user.groups.filter(name="Administrador").exists()

    if request.user != comunidad.creador and not es_admin:
        return HttpResponseForbidden()

    comunidad.delete()
    return redirect('comunidades')

# =====================================
# INVENTARIO ADMIN
# =====================================
@login_required



def inventario_admin(request):
    productos = Producto.objects.all()

    query = request.GET.get("q", "")
    filtro = request.GET.get("filtro", "")

    # 🔎 BUSCADOR
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(descripcion__icontains=query)
        )

    # ⚠️ FILTROS
    if filtro == "stock_bajo":
        productos = productos.filter(cantidad__lt=5, cantidad__gt=0)

    elif filtro == "sin_stock":
        productos = productos.filter(cantidad=0)

    # 📄 PAGINACIÓN
    paginator = Paginator(productos, 10)
    page = request.GET.get("page")
    productos = paginator.get_page(page)

    return render(request, "admin/inventario.html", {
        "productos": productos,
        "query": query,
        "filtro": filtro
    })

#======================================
#exportar pdf de productos del admin
#=======================================
@login_required
def exportar_pdf_inventario(request):
    q = request.GET.get("q", "")
    filtro = request.GET.get("filtro", "")
    productos = Producto.objects.all()
    if q:
        productos = productos.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
    if filtro == "stock_bajo":
        productos = productos.filter(cantidad__lt=5, cantidad__gt=0)
    elif filtro == "sin_stock":
        productos = productos.filter(cantidad=0)

    template = get_template("pdf_inventario.html")
    html = template.render({"productos": productos})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

#======================================
#exportar excel de productos del admin
#=======================================
@login_required
def exportar_excel_inventario(request):
    q = request.GET.get("q", "")
    filtro = request.GET.get("filtro", "")
    productos = Producto.objects.all()
    if q:
        productos = productos.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
    if filtro == "stock_bajo":
        productos = productos.filter(cantidad__lt=5, cantidad__gt=0)
    elif filtro == "sin_stock":
        productos = productos.filter(cantidad=0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.append(["ID", "Producto", "Descripción", "Precio", "Stock", "Vendedor"])

    for p in productos:
        ws.append([
            p.id,
            p.nombre,
            p.descripcion,
            float(p.precio),
            p.cantidad,
            p.vendedor.username
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=inventario.xlsx'

    wb.save(response)
    return response


# =====================================
# ELIMINAR PRODUCTO (ADMIN)
# =====================================
@login_required
def eliminar_producto_admin(request, id):

    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("redireccion_dashboard")

    producto = get_object_or_404(Producto, id=id)
    producto.delete()

    messages.success(request, "Producto eliminado correctamente.")

    return redirect("inventario_admin")
# =====================================
# CREAR USUARIO (ADMIN)
# =====================================
@login_required
def crear_usuario(request):

    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("redireccion_dashboard")

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        rol = request.POST.get("rol")

        if User.objects.filter(username=username).exists():
            messages.error(request, "El usuario ya existe.")
            return redirect("crear_usuario")

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        grupo, _ = Group.objects.get_or_create(name=rol)
        user.groups.add(grupo)

        messages.success(request, "Usuario creado correctamente.")
        return redirect("dashboard_admin")

    grupos = Group.objects.all()
    return render(request, "crear_usuario.html", {"grupos": grupos})


# =====================================
# EDITAR USUARIO (ADMIN)
# =====================================
@login_required
def editar_usuario(request, user_id):

    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("redireccion_dashboard")

    usuario = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        nuevo_username = request.POST.get("username") # 👈 Capturamos el nombre primero
        usuario.email = request.POST.get("email")

        rol = request.POST.get("rol")
        nueva_password = request.POST.get("password")

        # 🔥 VALIDACIÓN: Verificamos si ese username ya existe en OTRO usuario distinto
        if User.objects.filter(username=nuevo_username).exclude(id=user_id).exists():
            grupos = Group.objects.all()
            messages.error(request, f"El nombre de usuario '{nuevo_username}' ya está en uso por otra cuenta.")
            return render(request, "editar_usuario.html", {
                "usuario": usuario,
                "grupos": grupos
            })

        # Si el nombre no está repetido, se lo asignamos al usuario seguro
        usuario.username = nuevo_username

        if nueva_password:
            usuario.set_password(nueva_password)

        usuario.save() # 👈 Ya no romperá la base de datos

        usuario.groups.clear()
        grupo, _ = Group.objects.get_or_create(name=rol)
        usuario.groups.add(grupo)

        # Sincronizar Perfil.rol con el grupo asignado
        try:
            usuario.perfil.rol = rol
            usuario.perfil.save()
        except Exception:
            pass

        messages.success(request, "Usuario actualizado correctamente.")
        return redirect("dashboard_admin")

    grupos = Group.objects.all()

    return render(request, "editar_usuario.html", {
        "usuario": usuario,
        "grupos": grupos
    })
# =====================================
# exportar pdf de usuarios (ADMIN)
# =====================================
def exportar_pdf_usuarios(request):
    query = request.GET.get("q", "")
    rol = request.GET.get("rol", "")

    perfiles = Perfil.objects.select_related("user").all()
    if query:
        perfiles = perfiles.filter(
            Q(user__username__icontains=query) |
            Q(user__email__icontains=query)
        )
    if rol:
        perfiles = perfiles.filter(rol=rol)

    template = get_template("pdf_usuarios.html")
    html = template.render({"perfiles": perfiles})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="usuarios.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response

# =====================================
# exportar excel de usuarios (ADMIN)
# =====================================
def exportar_excel_usuarios(request):
    query = request.GET.get("q", "")
    rol = request.GET.get("rol", "")

    perfiles = Perfil.objects.select_related("user").all()
    if query:
        perfiles = perfiles.filter(
            Q(user__username__icontains=query) |
            Q(user__email__icontains=query)
        )
    if rol:
        perfiles = perfiles.filter(rol=rol)

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    ws.append(["Usuario", "Email", "Rol"])

    for perfil in perfiles:
        ws.append([
            perfil.user.username,
            perfil.user.email,
            perfil.rol
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'

    wb.save(response)

    return response



# =====================================
# ELIMINAR USUARIO (ADMIN)
# =====================================
@login_required
def eliminar_usuario(request, user_id):
    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("dashboard_admin")

    usuario = get_object_or_404(User, id=user_id)

    if usuario == request.user:
        messages.error(request, "No puedes eliminarte a ti mismo.")
        return redirect("dashboard_admin")

    usuario.delete()
    messages.success(request, f"Usuario eliminado correctamente.")
    return redirect("dashboard_admin")

# =====================================
# DASHBOARD VENDEDOR
# =====================================
@login_required
@never_cache


def dashboard_vendedor(request):

    if not request.user.groups.filter(name="Vendedor").exists():
        messages.error(request, "No tienes permiso.")
        return redirect("redireccion_dashboard")

    productos = Producto.objects.filter(vendedor=request.user).order_by('-id')

    # 🔎 BUSQUEDA
    q = request.GET.get('q')
    if q:
        productos = productos.filter(nombre__icontains=q)

    # 🔥 CONTADORES (ANTES de paginar)
    total_productos = productos.count()
    publicados = productos.filter(publicado=True).count()
    borradores = productos.filter(publicado=False).count()
    activos = productos.filter(cantidad__gt=1).count()

    # 📄 PAGINADOR (5 por página)
    paginator = Paginator(productos, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ⚡ AJAX (clave para que funcione tu dashboard pro)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, "partials/tabla_productos.html", {
            "page_obj": page_obj
        })

    return render(request, "dashboard_vendedor.html", {
        "page_obj": page_obj,   # 👈 IMPORTANTE
        "total_productos": total_productos,
        "publicados": publicados,
        "borradores": borradores,
        "activos": activos
    })





# =====================================
# pdf de productos publicados del vendedor
# =====================================
@login_required
def exportar_pdf_vendedor(request):

    q = request.GET.get("q", "")
    productos = Producto.objects.filter(vendedor=request.user)
    if q:
        productos = productos.filter(nombre__icontains=q)

    template = get_template("pdf_productos.html")
    html = template.render({"productos": productos, "usuario": request.user})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response
from openpyxl import Workbook

# =====================================
# excel de productos publicados del vendedor
# =====================================

@login_required
def exportar_excel_vendedor(request):

    q = request.GET.get("q", "")
    productos = Producto.objects.filter(vendedor=request.user)
    if q:
        productos = productos.filter(nombre__icontains=q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    ws.append(["Nombre", "Precio", "Cantidad", "Publicado"])
    for p in productos:
        ws.append([p.nombre, float(p.precio), p.cantidad, "Sí" if p.publicado else "No"])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=productos.xlsx"
    wb.save(response)
    return response



# =====================================
# CREAR PRODUCTO
# =====================================
@login_required

def crear_producto(request):

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        descripcion = request.POST.get("descripcion", "").strip()
        precio = request.POST.get("precio", "").strip()
        cantidad = request.POST.get("cantidad", "").strip()
        imagen = request.FILES.get("imagen")

        if not nombre or not descripcion or not precio or not cantidad or not imagen:
            messages.error(request, "Todos los campos son obligatorios, incluyendo la imagen.")
            return redirect("crear_producto")

        Producto.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            precio=precio,
            cantidad=cantidad,
            imagen=imagen,
            vendedor=request.user
        )

        messages.success(request, f"✅ Producto '{nombre}' creado correctamente.")
        return redirect("dashboard_vendedor")

        messages.success(request, "Producto creado correctamente.")
        return redirect("dashboard_vendedor")

    return render(request, "crear_producto.html")

from django.shortcuts import get_object_or_404, redirect

# =====================================
# destacar producto (Vendedor)
# =====================================

def toggle_destacado(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    producto.destacado = not producto.destacado
    producto.save()

    return redirect('dashboard_vendedor')  # o donde estés mostrando productos

# =====================================
# EDITAR PRODUCTO
# =====================================
@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    # Solo el dueño puede editar
    if producto.vendedor != request.user:
        return HttpResponseForbidden("No tienes permiso.")

    if request.method == "POST":

        nombre = request.POST.get("nombre", "").strip()
        descripcion = request.POST.get("descripcion", "").strip()
        precio = request.POST.get("precio", "").strip()
        cantidad = request.POST.get("cantidad", "").strip()

        # Validar campos obligatorios
        if not nombre or not descripcion or not precio or not cantidad:
            messages.error(request, "Todos los campos son obligatorios excepto la imagen.")
            return render(request, "editar_producto.html", {
                "producto": producto
            })

        producto.nombre = nombre
        producto.descripcion = descripcion

        try:
            producto.precio = float(precio)
            producto.cantidad = int(cantidad)
        except ValueError:
            messages.error(request, "Precio y cantidad deben ser números válidos.")
            return render(request, "editar_producto.html", {
                "producto": producto
            })

        if request.FILES.get("imagen"):
            producto.imagen = request.FILES.get("imagen")

        producto.save()

        messages.success(request, "Producto actualizado correctamente.")
        return redirect("dashboard_vendedor")

    return render(request, "editar_producto.html", {
        "producto": producto
    })

# =====================================
# ELIMINAR PRODUCTO
# =====================================
@login_required
def eliminar_producto_vendedor(request, producto_id):

    producto = get_object_or_404(Producto, id=producto_id)

    if producto.vendedor != request.user:
        messages.error(request, "No tienes permiso para eliminar este producto.")
        return redirect("dashboard_vendedor")

    producto.delete()

    messages.success(request, "Producto eliminado.")
    return redirect("dashboard_vendedor")
# =====================================
# PUBLICAR / DESPUBLICAR PRODUCTO
# =====================================
@login_required
def toggle_publicacion(request, producto_id):

    producto = get_object_or_404(Producto, id=producto_id)

    # 🔐 Validación: solo el dueño puede cambiarlo
    if producto.vendedor != request.user:
        messages.error(request, "No tienes permiso para modificar este producto.")
        return redirect("dashboard_vendedor")

    # 🔄 Alternar estado
    producto.publicado = not producto.publicado
    producto.save()

    estado = "publicado" if producto.publicado else "ocultado"
    messages.success(request, f"Producto {estado} correctamente.")

    return redirect("dashboard_vendedor")


# =====================================
# REDIRECCIÓN
# =====================================
@login_required
def redireccion_dashboard(request):

    if request.user.groups.filter(name="Administrador").exists():
        return redirect("dashboard_admin")
    elif request.user.groups.filter(name="Vendedor").exists():
        return redirect("dashboard_vendedor")
    else:
        return redirect("dashboard_usuario")


# =====================================
# LOGOUT
# =====================================
@login_required
def cerrar_sesion(request):
    # Guardar carrito en la base de datos antes de cerrar sesión
    if request.user.is_authenticated:
        try:
            clave = get_carrito_key(request)
            carrito = request.session.get(clave, {})
            request.user.perfil.carrito_guardado = json.dumps(carrito)
            request.user.perfil.save()
        except Exception:
            pass
    logout(request)
    return redirect("login")

def obtener_rol_seguro(user):
    # Primero intentamos con tu lógica actual (perfil)
    try:
        if user.is_authenticated and hasattr(user, 'perfil'):
            return user.perfil.rol
    except:
        pass
    
    # Si falla, el sistema de grupos nos dice la verdad
    if user.groups.filter(name="Administrador").exists():
        return "Administrador"
    if user.groups.filter(name="Vendedor").exists():
        return "Vendedor"
    return "Usuario"

@login_required
def gestionar_productos_admin(request):
    # Solo administradores pueden entrar
    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "Acceso denegado.")
        return redirect("inicio")

    productos = Producto.objects.all().order_by('-creado')
    
    # Buscador opcional
    query = request.GET.get("q")
    if query:
        productos = productos.filter(nombre__icontains=query)

    return render(request, "admin/gestionar_productos.html", {
        "productos": productos
    })

@login_required
def admin_editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    
    # Validar que sea admin
    if not request.user.groups.filter(name="Administrador").exists():
        return HttpResponseForbidden("No eres administrador.")

    if request.method == "POST":
        # Aquí permites editar cualquier campo
        producto.nombre = request.POST.get("nombre")
        producto.precio = request.POST.get("precio")
        producto.publicado = request.POST.get("publicado") == 'on' # Checkbox
        producto.save()
        messages.success(request, f"Producto '{producto.nombre}' editado por Admin.")
        return redirect("gestionar_productos_admin")
        
    return render(request, "admin/editar_producto.html", {"producto": producto})

@login_required
def admin_eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    
    if request.user.groups.filter(name="Administrador").exists():
        producto.delete()
        messages.success(request, "Producto eliminado por el Administrador.")
    else:
        messages.error(request, "No tienes permiso.")
        
    return redirect("gestionar_productos_admin")